"""
خدمة تحليل العطاءات القوية: retry، معالجة أخطاء، جودة النتائج، ودمج الـ chunks.
"""
import asyncio
import json
import logging
import re
from datetime import datetime, timezone
from enum import Enum
from typing import Dict, Any, List, Optional

from sqlalchemy.orm import Session

from app.models.rfp_analysis import RFPAnalysis, AnalysisStatus
from app.services.advanced_chunking import AdvancedChunker
from app.services.document_processor import DocumentProcessor
from app.services.ai_provider import get_ai_provider

logger = logging.getLogger("ataeru.rfp")

class QualityScore(str, Enum):
    EXCELLENT = "excellent"   # 95-100%
    GOOD = "good"             # 80-94%
    ACCEPTABLE = "acceptable" # 60-79%
    POOR = "poor"             # 40-59%
    FAILED = "failed"         # < 40%

CHUNK_ANALYSIS_PROMPT = """أنت متخصص في تحليل وثائق العطاءات.

قم بتحليل النص التالي واستخرج:
1. المتطلبات الإجبارية (mandatory_requirements)
2. الآجال النهائية (deadline)
3. معايير التقييم (evaluation_criteria)
4. الأسئلة الفنية (technical_questions)
5. ملخص جدول الكميات (boq_summary)

النص:
{text}

أعد الرد بصيغة JSON صارمة. إذا لم تجد معلومة، اترك القيمة null.
"""

class RobustRFPService:
    MAX_RETRIES = 3
    RETRY_DELAY = 2  # seconds
    CHUNK_TIMEOUT = 30  # seconds

    @staticmethod
    async def extract_and_analyze_robust(
        db: Session,
        analysis_id: int,
    ) -> Dict[str, Any]:
        """تحليل شامل مع retry ومعالجة أخطاء."""
        analysis = db.query(RFPAnalysis).filter(RFPAnalysis.id == analysis_id).first()
        if not analysis:
            logger.error(f"Analysis {analysis_id} not found")
            raise ValueError("Analysis not found")

        attempt_logs: List[Dict[str, Any]] = []
        max_retries = analysis.max_retries or RobustRFPService.MAX_RETRIES

        for attempt in range(max_retries):
            try:
                logger.info(f"Starting analysis {analysis_id}, attempt {attempt + 1}/{max_retries}")

                # 1. تحديث الحالة
                analysis.status = AnalysisStatus.PROCESSING
                analysis.current_attempt = attempt + 1
                analysis.progress = 0
                db.commit()

                # 2. استخراج النص
                extracted_text = RobustRFPService.safe_extract_text(
                    analysis.original_file_url,
                    analysis.original_file_name
                )

                if not extracted_text or len(extracted_text.strip()) < 100:
                    raise ValueError("Extracted text is too short or empty")

                analysis.extracted_text = extracted_text
                db.commit()
                logger.info(f"Text extracted: {len(extracted_text)} characters")

                # 3. تقسيم ذكي وفرز حسب الأهمية
                chunks = AdvancedChunker.smart_chunk(extracted_text)
                chunks = AdvancedChunker.sort_by_importance(chunks)
                logger.info(f"Text split into {len(chunks)} chunks")

                # 4. تحليل كل chunk
                chunk_results = []
                for i, chunk in enumerate(chunks):
                    logger.info(f"Analyzing chunk {i + 1}/{len(chunks)}")
                    chunk_result = await RobustRFPService.analyze_single_chunk(chunk)
                    if chunk_result:
                        chunk_results.append(chunk_result)
                    analysis.progress = int((i + 1) / len(chunks) * 100)
                    db.commit()

                if not chunk_results:
                    raise ValueError("No chunks were analyzed successfully")

                # 5. دمج النتائج
                merged_result = AdvancedChunker.merge_chunk_results(chunk_results)

                # 6. حساب جودة التحليل
                quality_score = RobustRFPService.calculate_quality_score(merged_result, extracted_text)

                # 7. التحقق من الجودة - إعادة المحاولة عند الجودة المنخفضة
                if quality_score in (QualityScore.FAILED, QualityScore.POOR):
                    logger.warning(f"Low quality analysis: {quality_score.value}")
                    if attempt < max_retries - 1:
                        attempt_logs.append({
                            "attempt": attempt + 1,
                            "status": "low_quality",
                            "quality_score": quality_score.value,
                        })
                        await asyncio.sleep(RobustRFPService.RETRY_DELAY)
                        continue

                # 8. نجاح التحليل
                RobustRFPService._save_result(analysis, merged_result, quality_score, attempt, attempt_logs)
                db.commit()

                logger.info(
                    f"Analysis {analysis_id} completed with quality: {quality_score.value}, "
                    f"chunks: {len(chunk_results)}"
                )

                return {
                    "success": True,
                    "analysis_id": analysis_id,
                    "quality_score": quality_score.value,
                    "confidence": analysis.confidence_score,
                    "result": merged_result,
                    "attempts": attempt + 1,
                    "chunks_analyzed": len(chunk_results),
                }

            except Exception as e:
                error_msg = f"Attempt {attempt + 1} failed: {str(e)}"
                logger.error(error_msg, exc_info=True)

                attempt_logs.append({
                    "attempt": attempt + 1,
                    "status": "failed",
                    "error": str(e),
                })

                if attempt == max_retries - 1:
                    return RobustRFPService._save_failure(analysis, error_msg, attempt_logs, db)

                await asyncio.sleep(RobustRFPService.RETRY_DELAY * (attempt + 1))

        # دفاع: لا يجب الوصول هنا عادة
        return RobustRFPService._save_failure(analysis, "Max retries exceeded", attempt_logs, db)

    @staticmethod
    def _save_result(
        analysis: RFPAnalysis,
        merged_result: Dict[str, Any],
        quality_score: QualityScore,
        attempt: int,
        attempt_logs: List[Dict[str, Any]],
    ):
        """حفظ النتيجة مع مزامنة الحقول القديمة للتوافق."""
        analysis.analysis_result = merged_result
        analysis.extracted_requirements = merged_result
        analysis.mandatory_checklist = merged_result.get("mandatory_requirements", [])
        analysis.evaluation_criteria = merged_result.get("evaluation_criteria", {}) or {}
        analysis.quality_score = quality_score.value
        analysis.status = AnalysisStatus.COMPLETED
        analysis.completed_at = datetime.now(timezone.utc)
        analysis.confidence_score = RobustRFPService.calculate_confidence(quality_score)
        analysis.attempt_logs = attempt_logs
        analysis.retry_count = attempt
        analysis.error_message = None
        analysis.progress = 100

    @staticmethod
    def _save_failure(
        analysis: RFPAnalysis,
        error_msg: str,
        attempt_logs: List[Dict[str, Any]],
        db: Session,
    ) -> Dict[str, Any]:
        """تسجيل الفشل النهائي وإخطار المستخدم عبر الحالة والسجلات."""
        logger.error(f"All attempts failed for analysis {analysis.id}: {error_msg}")
        analysis.status = AnalysisStatus.FAILED
        analysis.error_message = error_msg
        analysis.quality_score = QualityScore.FAILED.value
        analysis.confidence_score = 0
        analysis.attempt_logs = attempt_logs
        analysis.completed_at = datetime.now(timezone.utc)
        db.commit()

        RobustRFPService.notify_user_of_failure(analysis.id, error_msg)

        return {
            "success": False,
            "analysis_id": analysis.id,
            "status": AnalysisStatus.FAILED,
            "error": error_msg,
            "attempts": len(attempt_logs),
        }

    @staticmethod
    def safe_extract_text(file_path: str, file_name: str) -> str:
        """استخراج آمن للنص مع معالجة الأخطاء."""
        try:
            file_type = file_name.rsplit('.', 1)[-1].upper() if '.' in file_name else ""

            if file_type == "PDF":
                text = DocumentProcessor._extract_text_from_pdf(file_path)
                if not text.strip():
                    raise ValueError("لم يتم استخراج أي نص من الملف")
                return text

            if file_type in ("DOCX", "DOC"):
                text = DocumentProcessor._extract_text_from_docx(file_path)
                if not text.strip():
                    raise ValueError("لم يتم استخراج أي نص من الملف")
                return text

            if file_type in ("XLSX", "XLS"):
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    text += f"Sheet: {sheet}\n"
                    for row in ws.iter_rows(values_only=True):
                        text += " | ".join(str(cell) for cell in row if cell is not None) + "\n"
                wb.close()
                return text.strip()

            if file_type == "TXT":
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()

            raise ValueError(f"Unsupported file type: {file_type or 'unknown'}")

        except FileNotFoundError:
            raise ValueError(f"File not found: {file_path}")
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"Error extracting text: {str(e)}")

    @staticmethod
    async def analyze_single_chunk(chunk: Dict) -> Optional[Dict]:
        """تحليل قطعة واحدة مع معالجة أخطاء واسترجاع JSON."""
        try:
            ai_provider = get_ai_provider()
            prompt = CHUNK_ANALYSIS_PROMPT.format(text=chunk["text"])

            messages = [
                {"role": "system", "content": "أنت محلل متخصص في وثائق العطاءات. أعد JSON صارم فقط بدون تعليقات."},
                {"role": "user", "content": prompt},
            ]

            response = await asyncio.wait_for(
                asyncio.to_thread(
                    ai_provider.generate_completion,
                    messages,
                    0.2,
                    2000,
                ),
                timeout=RobustRFPService.CHUNK_TIMEOUT,
            )

            cleaned = response.strip()
            if cleaned.startswith("```json"):
                cleaned = cleaned[7:]
            if cleaned.endswith("```"):
                cleaned = cleaned[:-3]

            try:
                result = json.loads(cleaned)
                if isinstance(result, dict):
                    return result
            except json.JSONDecodeError as e:
                logger.warning(f"JSON decode error: {e}, attempting recovery")

            # استرجاع JSON من النص
            json_match = re.search(r'\{.*\}', cleaned, re.DOTALL)
            if json_match:
                try:
                    result = json.loads(json_match.group())
                    if isinstance(result, dict):
                        return result
                except (json.JSONDecodeError, AttributeError):
                    pass

            logger.warning("Failed to parse chunk result as JSON")
            return None

        except asyncio.TimeoutError:
            logger.error("Chunk analysis timeout")
            return None
        except Exception as e:
            logger.error(f"Error analyzing chunk: {str(e)}")
            return None

    @staticmethod
    def calculate_quality_score(result: Dict, original_text: str) -> QualityScore:
        """حساب درجة جودة التحليل (0-100)."""
        score = 0
        total = 0

        # 1. المتطلبات (25%)
        if "mandatory_requirements" in result:
            reqs_count = len(result.get("mandatory_requirements", []))
            if reqs_count > 0:
                score += min(reqs_count / 5, 1) * 25
            total += 25

        # 2. الآجال (20%)
        if "deadline" in result:
            if result.get("deadline"):
                score += 20
            total += 20

        # 3. معايير التقييم (20%)
        if "evaluation_criteria" in result:
            if result.get("evaluation_criteria"):
                score += 20
            total += 20

        # 4. الأسئلة الفنية (20%)
        if "technical_questions" in result:
            q_count = len(result.get("technical_questions", []))
            if q_count > 0:
                score += min(q_count / 3, 1) * 20
            total += 20

        # 5. BOQ (15%)
        if "boq_summary" in result:
            if result.get("boq_summary"):
                score += 15
            total += 15

        percentage = (score / total * 100) if total > 0 else 0

        if percentage >= 95:
            return QualityScore.EXCELLENT
        elif percentage >= 80:
            return QualityScore.GOOD
        elif percentage >= 60:
            return QualityScore.ACCEPTABLE
        elif percentage >= 40:
            return QualityScore.POOR
        else:
            return QualityScore.FAILED

    @staticmethod
    def calculate_confidence(quality_score: QualityScore) -> int:
        """تحويل جودة التحليل إلى درجة ثقة (1-100)."""
        mapping = {
            QualityScore.EXCELLENT: 95,
            QualityScore.GOOD: 80,
            QualityScore.ACCEPTABLE: 65,
            QualityScore.POOR: 40,
            QualityScore.FAILED: 0,
        }
        return mapping.get(quality_score, 0)

    @staticmethod
    def notify_user_of_failure(analysis_id: int, error_msg: str):
        """إخطار المستخدم عند الفشل (يُسجَّل في السجلات والـ DB)."""
        logger.info(f"Notifying user about failed analysis {analysis_id}: {error_msg}")
